#!/usr/bin/env python
import sys,os
import openpyxl
from openpyxl import load_workbook



texto_a_buscar=sys.argv[1]

print sys.argv[1]

print texto_a_buscar

if len(sys.argv) < 2:
    print 'Usage: ' + sys.argv[0] + ' <filename>'
    sys.exit(1)

print 'This is the name of the python script: ' + sys.argv[0]
print 'This is the 1st argument:              ' + sys.argv[1]


workbook = load_workbook('/home/mik/scripts/hoja_excel_gonzalo.xlsx')
first_sheet = workbook.get_sheet_names()[0]
worksheet = workbook.get_sheet_by_name(first_sheet)

for index, row in enumerate(worksheet.iter_rows()):
    for cell in row:
        print(cell, cell.value)
        if cell.value==texto_a_buscar:
		print "ENCONTRADO!! en ", cell
