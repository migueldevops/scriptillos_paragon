#!/usr/bin/env python

from openpyxl import load_workbook

workbook = load_workbook('/home/mik/scripts/hoja_excel_gonzalo.xlsx')
first_sheet = workbook.get_sheet_names()[0]
worksheet = workbook.get_sheet_by_name(first_sheet)

for index, row in enumerate(worksheet.iter_rows()):
    for cell in row:
        print(cell, cell.value)
        if cell.value=='Ir al centro':
		print "IR AL CENTRO ENCONTRADO!!"
